#源码记录
# 该文件主要是获取NBA球员信息。
import requests
import openpyxl
from lxml import etree
from openpyxl.styles import Font
def get_nba():
    url = "https://nba.hupu.com/stats/players"
    headers = {'user-Agent' : 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/116.0.0.0 Safari/537.36'}
    # 得到的仅仅是网页HTML源代码:文本形式展示出来，不具有结构性
    res = requests.get(url,headers = headers)
    # 处理结果
    e = etree.HTML(res.text) #变为HTML网页代码，具有结构性
    # 使用XPath获取数据
    nos= e.xpath('//table[@class="players_table"]//tr/td[1]/text()')
    names = e.xpath('//table[@class="players_table"]//tr/td[2]/a/text()')
    teams = e.xpath('//table[@class="players_table"]//tr/td[3]/a/text()')
    scores = e.xpath('//table[@class="players_table"]//tr/td[4]/text()')
    # 数据遍历
    for no,name,team,score in zip(nos,names,teams,scores):
        print(f"排名：{no} 姓名：{name} 球队：{team} 得分：{score}\n")
    #数据保存
    # 创建一个新的Excel文件
    wb = openpyxl.Workbook()
    # 选择默认的活动工作表
    ws = wb.active
    # 设置表头样式为粗体
    bold_font = Font(bold=True)
    ws['A1'].font = bold_font
    ws['B1'].font = bold_font
    ws['C1'].font = bold_font
    ws['D1'].font = bold_font
    # 写入表头
    ws['A1'] = '排名'
    ws['B1'] = '姓名'
    ws['C1'] = '球队'
    ws['D1'] = '得分'
    # 写入数据
    for i, (no, name, team, score) in enumerate(zip(nos, names, teams, scores), start=2):
        ws.cell(row=i, column=1).value = no
        ws.cell(row=i, column=2).value = name
        ws.cell(row=i, column=3).value = team
        ws.cell(row=i, column=4).value = score
    # 保存Excel文件
    wb.save('data.xlsx')
# 按间距中的绿色按钮以运行脚本。
if __name__ == '__main__':
    get_nba()